from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import sqlite3, os, csv, io, json
import openpyxl
from functools import wraps
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side

app = Flask(__name__)
app.secret_key = "project_alloc_secret_2024"
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 5 * 1024 * 1024

DB = 'instance/project_allocation.db'
os.makedirs('uploads', exist_ok=True)
os.makedirs('instance', exist_ok=True)

# ─── DATABASE ───────────────────────────────────────────────────────────────
def get_db():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    c = conn.cursor()

    # Create all tables (fresh install)
    c.executescript('''
        CREATE TABLE IF NOT EXISTS coordinator (
            id INTEGER PRIMARY KEY,
            name TEXT NOT NULL,
            email TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS guide (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            email TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            department TEXT,
            expertise TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS student (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            prn TEXT UNIQUE NOT NULL,
            division TEXT,
            department TEXT,
            email TEXT UNIQUE,
            password TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS project_group (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            team_lead_id INTEGER,
            project_title TEXT,
            title_finalized INTEGER DEFAULT 0,
            guide_id INTEGER,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(team_lead_id) REFERENCES student(id),
            FOREIGN KEY(guide_id) REFERENCES guide(id)
        );
        CREATE TABLE IF NOT EXISTS group_member (
            group_id INTEGER,
            student_id INTEGER,
            PRIMARY KEY(group_id, student_id),
            FOREIGN KEY(group_id) REFERENCES project_group(id),
            FOREIGN KEY(student_id) REFERENCES student(id)
        );
        CREATE TABLE IF NOT EXISTS project_title (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            guide_id INTEGER NOT NULL,
            title TEXT NOT NULL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(guide_id) REFERENCES guide(id)
        );
    ''')

    # ── Auto-migrations ─────────────────────────────────────────────────────
    existing_cols = [row[1] for row in c.execute("PRAGMA table_info(student)").fetchall()]
    if 'roll_no' in existing_cols and 'prn' not in existing_cols:
        print("  [Migration] roll_no -> prn ...")
        c.executescript('''
            ALTER TABLE student RENAME TO student_old;
            CREATE TABLE student (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                prn TEXT UNIQUE NOT NULL,
                division TEXT,
                department TEXT,
                email TEXT UNIQUE,
                password TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            );
            INSERT INTO student (id, name, prn, department, email, password, created_at)
                SELECT id, name, roll_no, department, email, password, created_at FROM student_old;
            DROP TABLE student_old;
        ''')
    elif 'prn' in existing_cols and 'division' not in existing_cols:
        c.execute("ALTER TABLE student ADD COLUMN division TEXT")

    pg_cols = [row[1] for row in c.execute("PRAGMA table_info(project_group)").fetchall()]
    if 'title_finalized' not in pg_cols:
        c.execute("ALTER TABLE project_group ADD COLUMN title_finalized INTEGER DEFAULT 0")
        print("  [Migration] Added title_finalized column.")

    # Default coordinator
    existing = c.execute("SELECT id FROM coordinator WHERE email='coordinator@college.edu'").fetchone()
    if not existing:
        c.execute("INSERT INTO coordinator (name, email, password) VALUES (?, ?, ?)",
                  ("Admin Coordinator", "coordinator@college.edu",
                   generate_password_hash("coord123")))
    conn.commit()
    conn.close()

# ─── AUTH DECORATORS ────────────────────────────────────────────────────────
def login_required(role):
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if session.get('role') != role:
                flash('Please login to continue.', 'warning')
                return redirect(url_for('login'))
            return f(*args, **kwargs)
        return decorated
    return decorator

# ─── ROUTES: AUTH ───────────────────────────────────────────────────────────
@app.route('/')
def index():
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        role = request.form.get('role')
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '').strip()
        conn = get_db()

        if role == 'coordinator':
            user = conn.execute("SELECT * FROM coordinator WHERE email=?", (email,)).fetchone()
            if user and check_password_hash(user['password'], password):
                session['role'] = 'coordinator'
                session['user_id'] = user['id']
                session['user_name'] = user['name']
                conn.close()
                return redirect(url_for('coord_dashboard'))
            flash('Invalid coordinator credentials.', 'error')

        elif role == 'guide':
            user = conn.execute("SELECT * FROM guide WHERE email=?", (email,)).fetchone()
            if user and check_password_hash(user['password'], password):
                session['role'] = 'guide'
                session['user_id'] = user['id']
                session['user_name'] = user['name']
                conn.close()
                return redirect(url_for('guide_dashboard'))
            flash('Invalid guide credentials.', 'error')

        elif role == 'student':
            user = conn.execute("SELECT * FROM student WHERE email=?", (email,)).fetchone()
            if user and check_password_hash(user['password'], password):
                session['role'] = 'student'
                session['user_id'] = user['id']
                session['user_name'] = user['name']
                conn.close()
                return redirect(url_for('student_dashboard'))
            flash('Invalid student credentials. Use your Email as login ID and PRN as password.', 'error')

        conn.close()
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# ─── ROUTES: COORDINATOR ────────────────────────────────────────────────────
@app.route('/coordinator/dashboard')
@login_required('coordinator')
def coord_dashboard():
    conn = get_db()
    stats = {
        'students': conn.execute("SELECT COUNT(*) FROM student").fetchone()[0],
        'guides': conn.execute("SELECT COUNT(*) FROM guide").fetchone()[0],
        'groups': conn.execute("SELECT COUNT(*) FROM project_group").fetchone()[0],
        'allocated': conn.execute("SELECT COUNT(*) FROM project_group WHERE guide_id IS NOT NULL").fetchone()[0],
        'finalized': conn.execute("SELECT COUNT(*) FROM project_group WHERE title_finalized=1").fetchone()[0],
        'pending': conn.execute("SELECT COUNT(*) FROM project_group WHERE title_finalized=0 AND project_title IS NOT NULL AND project_title != ''").fetchone()[0],
    }
    recent_groups = conn.execute("""
        SELECT pg.id, pg.project_title, pg.guide_id, pg.title_finalized,
               s.name as lead_name,
               g.name as guide_name,
               (SELECT COUNT(*) FROM group_member WHERE group_id=pg.id) as member_count
        FROM project_group pg
        LEFT JOIN student s ON pg.team_lead_id = s.id
        LEFT JOIN guide g ON pg.guide_id = g.id
        ORDER BY pg.created_at DESC LIMIT 5
    """).fetchall()
    conn.close()
    return render_template('coordinator/dashboard.html', stats=stats, recent_groups=recent_groups)
from flask import send_file
from openpyxl.styles import Alignment, Font, Border, Side
import io
@app.route('/coordinator/students')
@login_required('coordinator')
def coord_students():
    conn = get_db()
    students = conn.execute("""
        SELECT s.*,
               CASE WHEN gm.student_id IS NOT NULL THEN 1 ELSE 0 END as in_group
        FROM student s
        LEFT JOIN group_member gm ON s.id = gm.student_id
        ORDER BY s.name
    """).fetchall()
    conn.close()
    return render_template('coordinator/students.html', students=students)

@app.route('/coordinator/export-project-details')
@login_required('coordinator')
def export_project_details():
    conn = get_db()
    
    # 1. Fetch all groups with their allocated guides
    groups = conn.execute("""
        SELECT pg.id, pg.project_title, g.name as guide_name
        FROM project_group pg
        LEFT JOIN guide g ON pg.guide_id = g.id
        ORDER BY pg.id ASC
    """).fetchall()

    # Create a new workbook and select active sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Project Details"

    # --- 2. SET UNIVERSITY HEADERS (Rows 1-3) ---
    ws.merge_cells('A1:I1')
    ws['A1'] = "Sandip University"
    
    ws.merge_cells('A2:I2')
    ws['A2'] = "School Of Computer Science and Engineering"
    
    ws.merge_cells('A3:I3')
    ws['A3'] = "Project Details Academic Year 2026-27"

    # Style university headers
    header_font = Font(name='Arial', bold=True, size=14)
    for row in range(1, 4):
        cell = ws.cell(row=row, column=1)
        cell.alignment = Alignment(horizontal='center')
        cell.font = header_font

    # --- 3. SET TABLE HEADERS (Row 5) ---
    columns = [
        "Group No.", "Sr.No.", "Name Of Student", "PRN", 
        "Project Guide Name", "Project Title", "Domain", 
        "Industry Name", "Industry contact person info"
    ]
    
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=5, column=col_num)
        cell.value = column_title
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))

    # --- 4. FILL DATA AND MERGE CELLS ---
    current_row = 6
    for grp in groups:
        # Get members for this specific group
        members = conn.execute("""
            SELECT name, prn FROM student 
            WHERE id IN (SELECT student_id FROM group_member WHERE group_id=?)
        """, (grp['id'],)).fetchall()
        
        if not members: continue

        group_start_row = current_row

        for i, member in enumerate(members):
            ws.cell(row=current_row, column=2).value = i + 1
            ws.cell(row=current_row, column=3).value = member['name']
            ws.cell(row=current_row, column=4).value = member['prn']
            
            # Apply borders to all columns (1 to 9)
            for c in range(1, 10):
                ws.cell(row=current_row, column=c).border = Border(
                    left=Side(style='thin'), right=Side(style='thin'), 
                    top=Side(style='thin'), bottom=Side(style='thin'))
            current_row += 1

        # Merge Group-level data (Group ID, Guide, Title) across the member rows
        end_row = current_row - 1
        
        group_info = {
            1: grp['id'], 
            5: grp['guide_name'] or "Not Allocated", 
            6: grp['project_title'] or "TBD"
        }
        for col, value in group_info.items():
            ws.merge_cells(start_row=group_start_row, start_column=col, end_row=end_row, end_column=col)
            cell = ws.cell(row=group_start_row, column=col)
            cell.value = value
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
# --- 5. AUTO-RESIZE COLUMNS ---
    for col in ws.columns:
        max_length = 0
        # Use the .column index of the first cell and convert it to a letter safely
        column_letter = get_column_letter(col[0].column) 
        
        for cell in col:
            try:
                # MergedCell objects don't have a value attribute in the same way
                # We only check cells that actually contain data
                if hasattr(cell, 'value') and cell.value:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            except:
                pass
        
        # Set adjusted width (Minimum 10, Maximum 50)
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 50)

    conn.close()
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, 
                     download_name="Project_Details_2026-27.xlsx", 
                     as_attachment=True, 
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/coordinator/students/delete/<int:sid>', methods=['POST'])
@login_required('coordinator')
def delete_student(sid):
    conn = get_db()
    conn.execute("DELETE FROM group_member WHERE student_id=?", (sid,))
    conn.execute("DELETE FROM student WHERE id=?", (sid,))
    conn.commit()
    conn.close()
    flash('Student removed.', 'success')
    return redirect(url_for('coord_students'))

@app.route('/coordinator/guides')
@login_required('coordinator')
def coord_guides():
    conn = get_db()
    guides = conn.execute("""
        SELECT g.*,
               (SELECT COUNT(*) FROM project_group pg WHERE pg.guide_id = g.id) as group_count,
               (SELECT COUNT(*) FROM project_title pt WHERE pt.guide_id = g.id) as title_count
        FROM guide g ORDER BY g.name
    """).fetchall()
    conn.close()
    return render_template('coordinator/guides.html', guides=guides)

@app.route('/coordinator/guides/add', methods=['POST'])
@login_required('coordinator')
def add_guide():
    name = request.form.get('name', '').strip()
    email = request.form.get('email', '').strip()
    password = request.form.get('password', '').strip()
    department = request.form.get('department', '').strip()
    expertise = request.form.get('expertise', '').strip()
    if not name or not email or not password:
        flash('Name, email and password are required.', 'error')
        return redirect(url_for('coord_guides'))
    conn = get_db()
    try:
        conn.execute(
            "INSERT INTO guide (name, email, password, department, expertise) VALUES (?,?,?,?,?)",
            (name, email, generate_password_hash(password), department, expertise)
        )
        conn.commit()
        flash(f'Guide {name} added successfully.', 'success')
    except sqlite3.IntegrityError:
        flash('Email already exists.', 'error')
    conn.close()
    return redirect(url_for('coord_guides'))

@app.route('/coordinator/guides/delete/<int:gid>', methods=['POST'])
@login_required('coordinator')
def delete_guide(gid):
    conn = get_db()
    conn.execute("DELETE FROM project_title WHERE guide_id=?", (gid,))
    conn.execute("UPDATE project_group SET guide_id=NULL WHERE guide_id=?", (gid,))
    conn.execute("DELETE FROM guide WHERE id=?", (gid,))
    conn.commit()
    conn.close()
    flash('Guide removed.', 'success')
    return redirect(url_for('coord_guides'))

@app.route('/coordinator/groups')
@login_required('coordinator')
def coord_groups():
    conn = get_db()
    groups = conn.execute("""
        SELECT pg.id, pg.project_title, pg.title_finalized, pg.guide_id,
               s.name as lead_name, s.prn as lead_roll,
               g.name as guide_name,
               (SELECT COUNT(*) FROM group_member WHERE group_id=pg.id) as member_count
        FROM project_group pg
        LEFT JOIN student s ON pg.team_lead_id = s.id
        LEFT JOIN guide g ON pg.guide_id = g.id
        ORDER BY pg.created_at DESC
    """).fetchall()
    guides = conn.execute("SELECT id, name, department FROM guide ORDER BY name").fetchall()

    groups_with_members = []
    for grp in groups:
        members = conn.execute("""
            SELECT s.name, s.prn FROM group_member gm
            JOIN student s ON gm.student_id = s.id
            WHERE gm.group_id=?
        """, (grp['id'],)).fetchall()
        groups_with_members.append({'group': grp, 'members': members})

    conn.close()
    return render_template('coordinator/groups.html', groups=groups_with_members, guides=guides)

@app.route('/coordinator/groups/allocate', methods=['POST'])
@login_required('coordinator')
def allocate_guide():
    group_id = request.form.get('group_id')
    guide_id = request.form.get('guide_id') or None
    conn = get_db()
    conn.execute("UPDATE project_group SET guide_id=? WHERE id=?", (guide_id, group_id))
    conn.commit()
    conn.close()
    flash('Guide allocated successfully.', 'success')
    return redirect(url_for('coord_groups'))

@app.route('/coordinator/groups/finalize-title', methods=['POST'])
@login_required('coordinator')
def finalize_title():
    group_id = request.form.get('group_id')
    conn = get_db()
    grp = conn.execute("SELECT project_title, title_finalized FROM project_group WHERE id=?", (group_id,)).fetchone()
    if not grp:
        flash('Group not found.', 'error')
        conn.close()
        return redirect(url_for('coord_groups'))
    if not grp['project_title']:
        flash('Cannot finalize — the group has not submitted a project title yet.', 'error')
        conn.close()
        return redirect(url_for('coord_groups'))
    # Toggle finalized
    new_status = 0 if grp['title_finalized'] else 1
    conn.execute("UPDATE project_group SET title_finalized=? WHERE id=?", (new_status, group_id))
    conn.commit()
    conn.close()
    if new_status:
        flash('✅ Project title finalized successfully.', 'success')
    else:
        flash('Project title un-finalized. Students can now update it.', 'warning')
    return redirect(url_for('coord_groups'))

@app.route('/coordinator/allocations')
@login_required('coordinator')
def coord_allocations():
    conn = get_db()
    guides = conn.execute("SELECT * FROM guide ORDER BY name").fetchall()
    guide_data = []
    for guide in guides:
        groups = conn.execute("""
            SELECT pg.id, pg.project_title, pg.title_finalized,
                   s.name as lead_name, s.prn as lead_roll,
                   (SELECT COUNT(*) FROM group_member WHERE group_id=pg.id) as member_count
            FROM project_group pg
            LEFT JOIN student s ON pg.team_lead_id = s.id
            WHERE pg.guide_id=?
            ORDER BY pg.created_at DESC
        """, (guide['id'],)).fetchall()
        guide_data.append({'guide': guide, 'groups': groups})
    unallocated = conn.execute("""
        SELECT pg.id, pg.project_title, pg.title_finalized,
               s.name as lead_name,
               (SELECT COUNT(*) FROM group_member WHERE group_id=pg.id) as member_count
        FROM project_group pg
        LEFT JOIN student s ON pg.team_lead_id = s.id
        WHERE pg.guide_id IS NULL
    """).fetchall()
    conn.close()
    return render_template('coordinator/allocations.html', guide_data=guide_data, unallocated=unallocated)

# ─── ROUTES: GUIDE ──────────────────────────────────────────────────────────
@app.route('/guide/dashboard')
@login_required('guide')
def guide_dashboard():
    conn = get_db()
    guide_id = session['user_id']
    groups = conn.execute("""
        SELECT pg.id, pg.project_title, pg.title_finalized,
               s.name as lead_name, s.prn as lead_roll,
               (SELECT COUNT(*) FROM group_member WHERE group_id=pg.id) as member_count
        FROM project_group pg
        LEFT JOIN student s ON pg.team_lead_id = s.id
        WHERE pg.guide_id=?
        ORDER BY pg.created_at DESC
    """, (guide_id,)).fetchall()
    titles = conn.execute("SELECT * FROM project_title WHERE guide_id=? ORDER BY created_at DESC", (guide_id,)).fetchall()
    guide = conn.execute("SELECT * FROM guide WHERE id=?", (guide_id,)).fetchone()
    conn.close()
    return render_template('guide/dashboard.html', groups=groups, titles=titles, guide=guide)

@app.route('/guide/groups')
@login_required('guide')
def guide_groups():
    conn = get_db()
    guide_id = session['user_id']
    groups_raw = conn.execute("""
        SELECT pg.id, pg.project_title, pg.title_finalized,
               s.name as lead_name, s.prn as lead_roll,
               pg.created_at
        FROM project_group pg
        LEFT JOIN student s ON pg.team_lead_id = s.id
        WHERE pg.guide_id=?
        ORDER BY pg.created_at DESC
    """, (guide_id,)).fetchall()
    groups = []
    for grp in groups_raw:
        members = conn.execute("""
            SELECT s.name, s.prn, s.email,
                   CASE WHEN pg.team_lead_id = s.id THEN 1 ELSE 0 END as is_lead
            FROM group_member gm
            JOIN student s ON gm.student_id = s.id
            JOIN project_group pg ON gm.group_id = pg.id
            WHERE gm.group_id=?
        """, (grp['id'],)).fetchall()
        groups.append({'group': grp, 'members': members})
    conn.close()
    return render_template('guide/groups.html', groups=groups)

@app.route('/guide/titles')
@login_required('guide')
def guide_titles():
    conn = get_db()
    titles = conn.execute("SELECT * FROM project_title WHERE guide_id=? ORDER BY created_at DESC", (session['user_id'],)).fetchall()
    conn.close()
    return render_template('guide/titles.html', titles=titles)

@app.route('/guide/titles/add', methods=['POST'])
@login_required('guide')
def add_title():
    title = request.form.get('title', '').strip()
    if not title:
        flash('Title cannot be empty.', 'error')
        return redirect(url_for('guide_titles'))
    conn = get_db()
    conn.execute("INSERT INTO project_title (guide_id, title) VALUES (?,?)", (session['user_id'], title))
    conn.commit()
    conn.close()
    flash('Project title added.', 'success')
    return redirect(url_for('guide_titles'))

@app.route('/guide/titles/delete/<int:tid>', methods=['POST'])
@login_required('guide')
def delete_title(tid):
    conn = get_db()
    conn.execute("DELETE FROM project_title WHERE id=? AND guide_id=?", (tid, session['user_id']))
    conn.commit()
    conn.close()
    flash('Title removed.', 'success')
    return redirect(url_for('guide_titles'))

@app.route('/guide/submissions')
@login_required('guide')
def guide_submissions():
    conn = get_db()
    groups = conn.execute("""
        SELECT pg.id, pg.project_title, pg.title_finalized, pg.created_at,
               s.name as lead_name, s.prn as lead_roll,
               (SELECT COUNT(*) FROM group_member WHERE group_id=pg.id) as member_count
        FROM project_group pg
        LEFT JOIN student s ON pg.team_lead_id = s.id
        WHERE pg.guide_id=? AND pg.project_title IS NOT NULL AND pg.project_title != ''
        ORDER BY pg.created_at DESC
    """, (session['user_id'],)).fetchall()
    conn.close()
    return render_template('guide/submissions.html', groups=groups)

# ─── ROUTES: STUDENT ────────────────────────────────────────────────────────
@app.route('/student/dashboard')
@login_required('student')
def student_dashboard():
    conn = get_db()
    sid = session['user_id']
    group = conn.execute("""
        SELECT pg.*, g.name as guide_name, g.email as guide_email, g.department as guide_dept,
               s.name as lead_name, s.prn as lead_roll
        FROM group_member gm
        JOIN project_group pg ON gm.group_id = pg.id
        LEFT JOIN guide g ON pg.guide_id = g.id
        LEFT JOIN student s ON pg.team_lead_id = s.id
        WHERE gm.student_id=?
    """, (sid,)).fetchone()

    members = []
    if group:
        members = conn.execute("""
            SELECT s.name, s.prn, s.email,
                   CASE WHEN pg.team_lead_id = s.id THEN 1 ELSE 0 END as is_lead
            FROM group_member gm
            JOIN student s ON gm.student_id = s.id
            JOIN project_group pg ON gm.group_id = pg.id
            WHERE gm.group_id=?
        """, (group['id'],)).fetchall()

    available = []
    if not group:
        available = conn.execute("""
            SELECT s.id, s.name, s.prn, s.division, s.department FROM student s
            WHERE s.id != ?
            AND s.id NOT IN (SELECT student_id FROM group_member)
            ORDER BY s.name
        """, (sid,)).fetchall()

    titles = conn.execute("""
        SELECT pt.id, pt.title, g.name as guide_name
        FROM project_title pt
        JOIN guide g ON pt.guide_id = g.id
        ORDER BY g.name, pt.title
    """).fetchall()

    student = conn.execute("SELECT * FROM student WHERE id=?", (sid,)).fetchone()
    conn.close()
    return render_template('student/dashboard.html',
                           group=group, members=members,
                           available=available, titles=titles,
                           student=student)

@app.route('/student/form-group', methods=['POST'])
@login_required('student')
def form_group():
    sid = session['user_id']
    conn = get_db()

    existing = conn.execute("SELECT group_id FROM group_member WHERE student_id=?", (sid,)).fetchone()
    if existing:
        flash('You are already in a group.', 'error')
        conn.close()
        return redirect(url_for('student_dashboard'))

    member_ids = request.form.getlist('members')
    team_lead = request.form.get('team_lead')
    title_type = request.form.get('title_type')
    project_title = request.form.get('project_title', '').strip()
    custom_title = request.form.get('custom_title', '').strip()

    final_title = custom_title if title_type == 'custom' else project_title
    all_members = list(set([str(sid)] + member_ids))

    if len(all_members) < 2:
        flash('Select at least 1 more member.', 'error')
        conn.close()
        return redirect(url_for('student_dashboard'))
    if len(all_members) > 4:
        flash('Maximum 4 members per group.', 'error')
        conn.close()
        return redirect(url_for('student_dashboard'))
    if not team_lead:
        flash('Please select a team lead.', 'error')
        conn.close()
        return redirect(url_for('student_dashboard'))

    for mid in all_members:
        in_grp = conn.execute("SELECT group_id FROM group_member WHERE student_id=?", (mid,)).fetchone()
        if in_grp:
            s = conn.execute("SELECT name FROM student WHERE id=?", (mid,)).fetchone()
            flash(f'{s["name"]} is already in a group.', 'error')
            conn.close()
            return redirect(url_for('student_dashboard'))

    pg = conn.execute(
        "INSERT INTO project_group (team_lead_id, project_title, title_finalized) VALUES (?,?,0)",
        (team_lead, final_title)
    )
    group_id = pg.lastrowid
    for mid in all_members:
        conn.execute("INSERT INTO group_member (group_id, student_id) VALUES (?,?)", (group_id, mid))
    conn.commit()
    conn.close()
    flash('Group formed successfully! 🎉', 'success')
    return redirect(url_for('student_dashboard'))

@app.route('/student/update-title', methods=['POST'])
@login_required('student')
def update_title():
    sid = session['user_id']
    conn = get_db()

    grp = conn.execute("""
        SELECT pg.id, pg.title_finalized FROM group_member gm
        JOIN project_group pg ON gm.group_id = pg.id
        WHERE gm.student_id=?
    """, (sid,)).fetchone()

    if not grp:
        flash('You are not in a group.', 'error')
        conn.close()
        return redirect(url_for('student_dashboard'))

    if grp['title_finalized']:
        flash('Your project title has been finalized by the coordinator and cannot be changed.', 'error')
        conn.close()
        return redirect(url_for('student_dashboard'))

    title_type = request.form.get('title_type')
    project_title = request.form.get('project_title', '').strip()
    custom_title = request.form.get('custom_title', '').strip()
    final_title = custom_title if title_type == 'custom' else project_title

    if not final_title:
        flash('Please select or enter a project title.', 'error')
        conn.close()
        return redirect(url_for('student_dashboard'))

    conn.execute("UPDATE project_group SET project_title=? WHERE id=?", (final_title, grp['id']))
    conn.commit()
    conn.close()
    flash('✅ Project title updated successfully.', 'success')
    return redirect(url_for('student_dashboard'))

# ─── RUN ────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    init_db()
    print("\n" + "="*55)
    print("  Project Allocation System - Running!")
    print("="*55)
    print("  URL: http://localhost:5500")
    print("  Coordinator: coordinator@college.edu / coord123")
    print("  Students: Email (login) / PRN (password) — after upload")
    print("="*55 + "\n")
    app.run(debug=True, port=5500)